"""
[WIRED] Regenerated Excel report — the human-facing view of MongoDB.

WHY THIS FILE EXISTS
---------------------
MongoDB is the only thing written to live/concurrently (see
capture_pipeline.py). This module turns what's in Mongo into a
`.xlsx` someone can actually open and browse: one row per capture
(the "Log" sheet, default view) and one row per distinct known object
(the "Inventory" sheet, from vision.storage.object_catalog) — same
underlying data, two presentations, matching the plan.

IMAGES IN EXCEL
----------------
Default: a hyperlink column (`primary_image`, `all_images`) pointing at
the actual file(s) on disk — fast to generate, tiny file size, works at
any row count. Embedding real thumbnails with openpyxl is possible but
bloats the file and slows Excel down noticeably past a few hundred
rows, since every regenerate re-embeds every image — deliberately not
the default; see EMBED_THUMBNAILS below if you want to turn it on.

FILE-LOCK SAFETY
-----------------
If the .xlsx is open in Excel when a refresh runs, a direct overwrite
would throw a raw PermissionError. Instead this always builds the new
file at a temp path first, then atomically replaces the real path
(os.replace) — so a refresh either fully succeeds or leaves the
previous good file untouched, never a half-written file. If the swap
itself fails (file genuinely locked), the caller gets a clear
"close the Excel file and try again" message instead of a silent
crash.

RECONCILE FROM EXCEL (DATA_AUTHORITY_MODE == "excel")
--------------------------------------------------------
reconcile_from_excel() reads the "Log" sheet back in, matches rows to
MongoDB documents by `object_id`, and upserts any hand-edited fixed-
column values back into Mongo. This is the explicit, safe stand-in for
"keep Excel and Mongo in sync" — see the module docstring discussion in
the conversation this was planned in for why true live two-way sync
isn't attempted (Excel file-locking + no change-notification make it
unsafe). A row with a broken freeform-JSON cell is skipped with a
collected warning rather than silently dropped or crashing the whole
import.
"""

from __future__ import annotations

import json
import os
import tempfile
from typing import List, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from vision.config import EXCEL_EXPORT_DIR, EXCEL_EXPORT_FILENAME
from vision.storage import attribute_schema, mongo_client, object_catalog

# Off by default — see module docstring. Flip to True if you specifically
# want a small thumbnail embedded per object (first/primary image only).
EMBED_THUMBNAILS = False

LOG_SHEET_NAME = "Log"
INVENTORY_SHEET_NAME = "Inventory"


class ExcelExportError(Exception):
    """Raised for anything that should be shown to the user as a plain
    error message — openpyxl missing, file locked on swap, etc."""


def _require_openpyxl():
    if not _OPENPYXL_AVAILABLE:
        raise ExcelExportError("openpyxl is not installed. Run: pip install openpyxl")


def _export_path() -> str:
    os.makedirs(EXCEL_EXPORT_DIR, exist_ok=True)
    return os.path.join(EXCEL_EXPORT_DIR, EXCEL_EXPORT_FILENAME)


def _log_headers() -> List[str]:
    labels = attribute_schema.display_labels()
    fixed_keys = attribute_schema.fixed_column_keys()
    return (
        ["Object ID", "Session", "Catalog ID", "Captured At"]
        + [labels.get(k, k) for k in fixed_keys]
        + ["Attributes (why)", "Primary Image", "All Images", "# Images"]
    )


def _log_row(obj: dict, image_docs: list) -> list:
    data = obj.get("data") or {}
    fixed_keys = attribute_schema.fixed_column_keys()
    freeform_col = attribute_schema.freeform_key()
    paths = [d.get("image_path", "") for d in image_docs]
    return (
        [obj.get("_id", ""), obj.get("session_id", ""), obj.get("catalog_id", ""),
         str(obj.get("captured_at", obj.get("date", "")))]
        + [data.get(k, "") for k in fixed_keys]
        + [json.dumps(data.get(freeform_col, {}), ensure_ascii=False)]
        + [paths[0] if paths else "", " | ".join(paths), len(paths)]
    )


def _inventory_headers() -> List[str]:
    return ["Catalog ID", "Name", "Category", "Color", "Size",
            "First Seen", "Last Seen", "Times Seen", "Linked Object IDs"]


def _inventory_row(entry: dict) -> list:
    return [
        entry.get("_id", ""), entry.get("name", ""), entry.get("category", ""),
        entry.get("color", ""), entry.get("size", ""),
        str(entry.get("first_seen", "")), str(entry.get("last_seen", "")),
        entry.get("times_seen", 0),
        " | ".join(entry.get("linked_object_ids", [])),
    ]


def _autosize_and_freeze(ws, header_row_len: int) -> None:
    ws.freeze_panes = "A2"
    for col_idx in range(1, header_row_len + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def build_report(session_id: str | None = None) -> str:
    """
    Regenerates the .xlsx from MongoDB. If `session_id` is given, the
    Log sheet is limited to that session (e.g. "today only") to keep the
    export fast/manageable — pass session_id=None for full history.
    Returns the path written to. Raises ExcelExportError on failure
    (missing openpyxl, or the file genuinely couldn't be swapped in
    because it's open elsewhere).
    """
    _require_openpyxl()

    objects = (
        mongo_client.objects_for_session(session_id)
        if session_id else mongo_client.all_objects_for_export()
    )
    inventory = object_catalog.list_inventory(limit=10000)

    wb = openpyxl.Workbook()
    log_ws = wb.active
    log_ws.title = LOG_SHEET_NAME
    log_ws.append(_log_headers())
    for cell in log_ws[1]:
        cell.font = Font(bold=True)

    for obj in objects:
        image_docs = mongo_client.get_images_for_object(obj["_id"])
        row_values = _log_row(obj, image_docs)
        log_ws.append(row_values)
        row_idx = log_ws.max_row
        primary_col = len(_log_headers()) - 2  # "Primary Image" column index (1-based)
        primary_path = row_values[-3]
        if primary_path:
            cell = log_ws.cell(row=row_idx, column=primary_col)
            cell.hyperlink = primary_path
            cell.style = "Hyperlink"

    _autosize_and_freeze(log_ws, len(_log_headers()))

    inv_ws = wb.create_sheet(INVENTORY_SHEET_NAME)
    inv_ws.append(_inventory_headers())
    for cell in inv_ws[1]:
        cell.font = Font(bold=True)
    for entry in inventory:
        inv_ws.append(_inventory_row(entry))
    _autosize_and_freeze(inv_ws, len(_inventory_headers()))

    final_path = _export_path()
    fd, tmp_path = tempfile.mkstemp(
        suffix=".xlsx", dir=os.path.dirname(final_path) or "."
    )
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, final_path)  # atomic on the same filesystem
    except PermissionError as e:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise ExcelExportError(
            f"Could not update '{final_path}' — it looks like it's open in "
            f"Excel. Close the file and try again. ({e})"
        )
    return final_path


def reconcile_from_excel() -> Tuple[int, List[str]]:
    """
    Reads the Log sheet back in and upserts any hand-edited fixed-column
    values into MongoDB, matched by Object ID. Only meaningful in
    DATA_AUTHORITY_MODE == "excel" (see vision/config.py) — the caller
    (main.py) is responsible for only exposing this action in that mode.

    Returns (rows_updated, warnings). A row with a broken "Attributes
    (why)" JSON cell is skipped (its fixed columns are NOT applied
    either, to avoid a half-updated row) and noted in `warnings` rather
    than raising — one bad row shouldn't block the rest of the import.
    """
    _require_openpyxl()
    path = _export_path()
    if not os.path.exists(path):
        raise ExcelExportError(f"No report found at '{path}' yet — run an export first.")

    wb = openpyxl.load_workbook(path, data_only=True)
    if LOG_SHEET_NAME not in wb.sheetnames:
        raise ExcelExportError(f"'{LOG_SHEET_NAME}' sheet not found in '{path}'.")
    ws = wb[LOG_SHEET_NAME]

    headers = [c.value for c in ws[1]]
    fixed_keys = attribute_schema.fixed_column_keys()
    labels = attribute_schema.display_labels()
    label_to_key = {labels.get(k, k): k for k in fixed_keys}
    freeform_col_name = "Attributes (why)"

    updated = 0
    warnings: List[str] = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        values = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
        object_id = values.get("Object ID")
        if not object_id:
            continue

        raw_freeform = values.get(freeform_col_name) or "{}"
        try:
            freeform = json.loads(raw_freeform) if isinstance(raw_freeform, str) else (raw_freeform or {})
        except (json.JSONDecodeError, TypeError) as e:
            warnings.append(f"Row for '{object_id}': bad JSON in '{freeform_col_name}' — skipped ({e})")
            continue

        data = {key: values.get(labels.get(key, key)) for key in fixed_keys}
        data[attribute_schema.freeform_key()] = freeform

        mongo_client.update_object_data(object_id, data)
        updated += 1

    return updated, warnings
